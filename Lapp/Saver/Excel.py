from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from EmailHandler.EmailProccessor import EmailsHandler
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
import string

class ExcelProcessor (EmailsHandler):
    __originalFile = []
    __fileProduced = None
    __sheetMade = []
    __error = None
    __isError = False
    __writeStartingRow = 0
    __dataProduced = []
    __acceptedDataFromFiles = []
    __sheetFoundInFile = []
    __collecttedData = []
    __rowsKeys = []
    __conditionArg = []
    __isEmail = False
    __fileCreated = False
    __progress = None

    def setAllRequired(self,originalFile =  [], fileProduce = None, sheetToMake = [], startWriteAt = 0):

        try:
            self.__originalFile = originalFile
            self.__fileProduced = fileProduce if fileProduce else "new_doc.xlsx"
            self.__sheetMade = sheetToMake
            self.__writeStartingRow = startWriteAt
        except:
            self.__isError = True
            self.__error = "Required more data"


    def setColumnKeys(self, rows = []):
        self.__rowsKeys = rows

    def setconditions(self, cond = []):
        self.__conditionArg = cond


    def setOriginalFile(self, filename = []):

        if(len(filename) != 0):
            self.__originalFile = filename
        else:
            self.__isError = True
            self.__error = "File name to read not set"


    def setProduceFile(self,filename = None):
        self.__fileProduced = filename if filename else "new_doc.xlsx"


    def setSheetName(self, sheets = []):
        self.__sheetMade = sheets

    def setWritingStartingRow(self, row= 0):
        self.__writeStartingRow = row

    def is_Error(self):
        return self.__isError

    def getError(self):
        return self.__error

    def getSortedData(self):
        return self.__dataProduced

    def getFileNameUploaded(self):
        return self.__originalFile

    def getSheetNames(self):
        return self.__sheetMade

    def savedFileName(self):
        return self.__fileProduced


    def readingAllFiles(self):

        try:
            for file in self.__originalFile:
                wb = load_workbook(file)
                ws = wb.active
                sheets = wb.sheetnames
                collection = {'sheet': sheets, 'path': file}
                self.__sheetFoundInFile.append(collection)
        except:
            self.__isError = True
            self.__error = "Something went wrong! in finding sheets"


    def getAllDataRead(self):
        return self.__acceptedDataFromFiles


    def getAllSheetFound(self):
        return self.__sheetFoundInFile


    def readActualData(self):

        all = []
        data = []

        try:
            for item in self.__sheetFoundInFile:
                path = item['path']
                sheets = item['sheet']
                for sheet in sheets:
                    wb = load_workbook(path)
                    ws = wb[sheet]
                    all.append({'sheet': sheet, 'data': ws})


                self.__collecttedData = all
        except Exception as e:
            self.__isError = True
            self.__error = "Something went wrong! read sheets data"
            print(e)


    def getActualData(self):
        return self.__collecttedData


    def processSheetByConditon(self):

        try:
            sizes = len(self.__collecttedData)
            for c in self.__conditionArg:
                for item in self.__collecttedData:
                    #self.__progress = self.__progress + "#" if self.__progress is not None else "#"
                    self.__proccesSheet(item['data'], item['sheet'], c.lower())

        except Exception as e:
            self.__isError = True
            self.__error = "Something went wrong! in processsheet condition"
            print(e)


    def __proccesSheet(self,sheet, sheetname, c):
        size = sheet.max_row
        i = 1
        try:
            while i < size:
                checked = sheet[i][self.__rowsKeys[1]]
                self.__row(sheet[i], checked.value, c)
                i += 1
        except Exception as e:
            self.__isError = True
            self.__error = "Something went wrong! in __prcessSheet method "
            print(e)


    def __row(self, colrow,checking, c):
        all = []
        subtring = "@"

        try:
            if checking is not None:
                if self.__isEmail == True:
                    if checking.find(subtring) != -1:
                        checking = checking.split('@')[-1].lower()
                    else:
                        checking = checking.lower()
                if (c.lower() == checking.lower()):
                    for item in colrow:
                        all.append(item.value)
                    self.writeFile(all, c)
        except:
            self.__isError = True
            self.__error = "Something went wrong in row data checking"


    def writeFile(self, datarow, sheetname):
        try:
            book = load_workbook(self.__fileProduced)
            sheetworkingon = book[sheetname]
            sheetworkingon.append(datarow)
            book.save(self.__fileProduced)

        except:
            self.__isError = True
            self.__error = "Something went wrong in writing row data"


    def createExcelSheet(self):

        try:
            wb = Workbook()

            if len(self.__conditionArg) != 0 and self.__fileCreated == False:
                for sheet in self.__conditionArg:
                    wb.create_sheet(sheet)
                wb.save(self.__fileProduced)
                self.__fileCreated = True
            else:
                print("Please set conditions first to be able to create file")
        except:
            self.__isError = True
            self.__error = "Something went wrong in creating saving file"


    def filter(self):

        try:
            self.getDataRequiredSheet(self.__conditionArg)
            emiallist = self.getEmailStmp()

            print("Do you want to filter by these mails: ", emiallist, "Enter (y or n)")
            confirmations = input()
            if confirmations.lower() == 'y':
                self.__isEmail = True
                self.setconditions(emiallist)
                self.createExcelSheet()
                self.writeFirstRow()
                print("processing please wait this will take some time to finish.....")
                self.processSheetByConditon()
                print("processing done saving.....")
                self.finalSave()
                self.__isEmail = False
            else:
                self.__isEmail = False
                self.setconditions(self.toLowerEmails(self.__conditionArg))
                self.createExcelSheet()
                self.writeFirstRow()
                print("processing please wait this will take some time to finish.....")
                self.processSheetByConditon()
                print("processing done saving.....")
                self.finalSave()
                self.__isEmail = False
        except NameError:
            self.__isError = True
            self.__error = NameError
        except Exception as e:
            self.__isError = True
            self.__error = "Something went wrong check the value you have entered!"
            print(e)


    def writeFirstRow(self, fields = []):

        try:
            inSheetFirstRowData = self.findFirstRow()

            if len(fields) == len(inSheetFirstRowData):
                self.createExcelSheet()
                for cond in self.__conditionArg:
                    book = load_workbook(self.__fileProduced)
                    sheetworkingon = book[cond]
                    sheetworkingon.append(fields)
                    book.save(self.__fileProduced)
            else:
                remark = input(
                    "You have enter less or more number of columns name than expected\nconfirm to proceed with old colums names (y or n): ")
                if remark.lower() == "y":
                    for cond in self.__conditionArg:
                        book = load_workbook(self.__fileProduced)
                        sheetworkingon = book[cond]
                        sheetworkingon.append(inSheetFirstRowData)
                        book.save(self.__fileProduced)
                else:
                    print("Operation cancelled!")
                    return
        except:
            self.__isError = True
            self.__error ="Something went wrong in writing first row"



    def findFirstRow(self):

        try:
            foundFirstData = []
            sheet = self.__collecttedData[0]['data']
            firstrow = None
            for row in sheet:
                firstrow = row
                break

            for col in firstrow:
                foundFirstData.append(col.value)
            return foundFirstData
        except:
            self.__isError = True
            self.__error = "Something went wrong! finding the row"


    def finalSave(self):
        print("\nsaving....")
        try:
            wb = load_workbook(self.__fileProduced)

            alpha = list(string.ascii_uppercase)
            font = Font(name="Arial Black", size=11, bold=True)
            background = PatternFill(start_color='EEEAE9', end_color='EEEAE9', fill_type='solid')
            align = Alignment(horizontal="center", vertical='center')

            i = 1

            for insideSheet in self.__conditionArg:
                ws = wb[insideSheet.lower()]
                dim_holder = DimensionHolder(worksheet=ws)

                for letter in alpha:
                    cell = ws[letter + str(i)]
                    cell.font = font
                    cell.fill = background
                    cell.alignment = align
                    wb.save(self.__fileProduced)
        except:
            self.__isError = True
            self.__error = "Something went wrong! save function"




